import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

f = open("opensource_cpp.txt", "r")

sheet['A1'] = "C++"
line = 'line'

while line != '':
    line = f.readline()
    translation_table = dict.fromkeys(map(ord, '\n\t'), None)
    target = line.translate(translation_table)
    target = target.split(':')
    if target[0] == 'open source':
        text = target[1]
        line = f.readline()
        translation_table = dict.fromkeys(map(ord, '\n\t'), None)
        target = line.translate(translation_table)
        target = target.split(':')
        text += " (" + target[1] + " )"
        if int(target[1]) == 0:
            break
        sheet.append([text])
    elif 'CVE' in target[0]:
        sheet.append(['', '', target[0], target[1]])
    elif '[' in target[0]:
        translation_table = dict.fromkeys(map(ord, '[]'), None)
        target = target[0].translate(translation_table)
        sheet.append(['',target])


    

f.close()

wb.save('test_oss_cpp.xlsx')

###########################################################################################################

wb = openpyxl.Workbook()
sheet = wb.active

f = open("cve_list_cpp.txt", "r")

sheet['A1'] = "C++"
line = 'line'

while line != '':
    line = f.readline()
    cves = []
    translation_table = dict.fromkeys(map(ord, '\n\t'), None)
    target = line.translate(translation_table)
    if target != '':
        line = f.readline()
        while ">" not in line:
            translation_table = dict.fromkeys(map(ord, '\n\t'), None)
            line = line.translate(translation_table)
            cves.append(line)
            line = f.readline()
        translation_table = dict.fromkeys(map(ord, '\n\t'), None)
        line = line.translate(translation_table)
        line = line.split(':')[1]
        target += " (" + line + " )"
        sheet.append([target])
        line = f.readline()
        for i in cves:
            sheet.append(['', i])



f.close()

wb.save('test_cve_cpp.xlsx')

########################################################################################################

wb = openpyxl.Workbook()
sheet = wb.active

f = open("links_cpp.txt", "r")

sheet['A1'] = "C++"
line = 'line'

ranking = 0

while line != '':
    ranking += 1
    line = f.readline()
    cves = []
    translation_table = dict.fromkeys(map(ord, '\n'), None)
    target = line.translate(translation_table)
    target = target.split('/')
    if len(target) == 5:
        sheet.append([ranking, target[3], target[4]])



f.close()

wb.save('ranking_cpp.xlsx')
